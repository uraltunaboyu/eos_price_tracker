from price_parser import PriceParser
import openpyxl
from openpyxl.worksheet.table import Table
from openpyxl.utils import column_index_from_string
from datetime import date

NAMES_FILE = "ItemLookUpTable_EN.lua"
PRICES_FILE = "PriceTableNA.lua"
CALCULATOR_FILE = "ESO_Crafting_Cost_Calculator.xlsx"

ALPHA = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
NUM = '0123456789'
COST_COL = 'Cost'

price_parser = PriceParser()
file = openpyxl.load_workbook(CALCULATOR_FILE)
file.save(CALCULATOR_FILE.split('.')[0] + "_old.xlsx")

debug_mode = True

def debug_print(*args):
    if debug_mode:
        print(*args)

def init_parser():
    price_parser.load_names(NAMES_FILE)
    price_parser.load_prices(PRICES_FILE)

def update_table(sheet_name: str, table_name: str = None, item_name_col: str = "Material"):
    debug_print(f"Using sheet {sheet_name}")
    if table_name is None:
        table_name = sheet_name
    else:
        debug_print(f"  Using table name {table_name}")
    sheet = file[sheet_name]
    table: Table = sheet.tables[table_name]
    table_dict = table_to_dict(table, sheet)
    for i, name in enumerate(table_dict[item_name_col]):
        debug_print(f"      Looking up {name.value}")
        sale_avg = price_parser.get_attr(name.value.lower(), 'SaleAvg')
        debug_print(f"      Found val {sale_avg}")
        if sale_avg is None:
            table_dict[COST_COL][1].value = ''
        else:
            table_dict[COST_COL][i].value = float(sale_avg)
    # cursed manual update for cell D3
    if table_name == "Styles":
        iii = float(price_parser.get_attr(table_dict[item_name_col][1].value.lower(), 'SaleAvg'))
        table_dict[COST_COL][1].value = iii
    update_timestamp(table, sheet)

def update_prices_chart():
    prices_chart_sheet = file['Price Charting']
    last_date = None
    last_col = None
    next_col = None

    for col in prices_chart_sheet.iter_cols(min_row=1, max_row=1, min_col=1):
        cell = col[0]
        last_date = cell.value
        last_col = cell.column

    debug_print(f"Last date is {last_date}")
    
    if last_date == date.today().isoformat():
        print("Chart does not need updating!")
        return
    
    next_col = last_col + 1
    timestamp_cell = prices_chart_sheet.cell(row = 1, column = next_col)
    debug_print(f"Updating prices for {date.today().isoformat()}")
    timestamp_cell.value = date.today().isoformat()
    timestamp_cell.number_format = "yyyy-mm-dd"
    
    for row in prices_chart_sheet.iter_rows(min_row=2, min_col=1, max_col=next_col):
        name_cell = row[0]
        target_cell = row[next_col - 1] 
        debug_print(f"  Updating price chart data for {name_cell.value}")
        try:
            sale_avg = price_parser.get_attr(name_cell.value.lower(), 'SaleAvg')
        except:
            continue
        if sale_avg is None:
            target_cell.value = ''
        else:
            target_cell.value = float(sale_avg)
            target_cell.number_format = "#,##0.00"

def update_timestamp(table: Table, sheet):
    timestamp = str(date.today())
    bottom_right: str = table.ref.split(':')[1]
    timestamp_row = int(bottom_right.lstrip(ALPHA)) + 1
    timestamp_col = column_index_from_string(bottom_right.rstrip(NUM))
    timestamp_cell = sheet.cell(row = timestamp_row, column = timestamp_col)
    timestamp_cell.value = timestamp
    timestamp_cell.number_format = "yyyy-mm-dd"

def table_to_dict(table: Table, sheet):
    top_left: str = table.ref.split(':')[0]
    bottom_right: str = table.ref.split(':')[1]

    start_col: str = column_index_from_string(top_left.rstrip(NUM))
    end_col: str = column_index_from_string(bottom_right.rstrip(NUM))
    start_row: int = int(top_left.lstrip(ALPHA))
    end_row: int = int(bottom_right.lstrip(ALPHA))

    data = {}

    header_row = sheet[start_row]
    for cell in header_row:
        if cell.column >= start_col and cell.column <= end_col:
            data[cell.value] = []

    for row in range(start_row + 1, end_row + 1):
        for col in range(start_col, end_col + 1):
            header = sheet.cell(row = start_row, column = col).value
            data[header].append(sheet.cell(row = row, column = col))

    return data

if __name__ == "__main__":
    init_parser()
    update_table("Blacksmithing")
    update_table("Clothing")
    update_table("Woodworking")
    update_table("Jeweling")
    update_table("Alchemy", item_name_col = "Reagent")
    update_table("Enchanting", table_name = "EnchantingPotency", item_name_col = "Potency")
    update_table("Enchanting", table_name = "EnchantingEssence", item_name_col = "Essence")
    update_table("Enchanting", table_name = "EnchantingAspect", item_name_col = "Aspect")
    update_table("Provisioning", table_name ="ProvisioningFood", item_name_col = "Ingredient")
    update_table("Provisioning", table_name = "ProvisioningDrink", item_name_col = "Ingredient")
    update_table("Provisioning", table_name = "ProvisioningOther", item_name_col = "Ingredient")
    update_table("Provisioning", table_name = "ProvisioningBait", item_name_col = "Ingredient")
    update_table("Furnishing")
    update_table("Style & Trait Materials", table_name = "Styles", item_name_col = "Style Material")
    update_table("Style & Trait Materials", table_name = "WeaponTraits")
    update_table("Style & Trait Materials", table_name = "ArmourTraits")
    update_table("Style & Trait Materials", table_name = "JewelryTraits")
    update_prices_chart()

    new_file_name = CALCULATOR_FILE
    file.save(new_file_name)